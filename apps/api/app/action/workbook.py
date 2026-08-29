"""Formatted Excel export (FR-011: "export results, narrative and metadata
to .xlsx"; docs/10_IMPLEMENTATION_ROADMAP.md Phase 7: "formatted workbook
with Summary, Data, SQL & Evidence, and Validation sheets").

Generated on demand from an already-validated RunSnapshot and returned as
bytes -- never written to disk (docs/09_DEPLOYMENT_OPERATIONS.md: "Excel
files generated on demand and downloaded immediately"). Deterministic
given the same snapshot, so a repeated download for the same
idempotency_key (app.action.store) produces a byte-for-byte-equivalent
workbook without needing to persist and re-serve a cached file.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.orchestrator.schema import AttemptRecord, RunSnapshot

_BOLD = Font(bold=True)


def _autosize(ws: Worksheet, *, max_width: int = 60) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = min(
                max_width, max(widths.get(cell.column, 0), len(str(cell.value)) + 2)
            )
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def _write_summary_sheet(ws: Worksheet, snapshot: RunSnapshot, latest: AttemptRecord) -> None:
    ws.title = "Summary"
    rows: list[tuple[str, object]] = [
        ("Question", snapshot.question),
        ("Run ID", str(snapshot.run_id)),
        ("Status", snapshot.status),
        ("Data as of", snapshot.completed_at.isoformat() if snapshot.completed_at else ""),
    ]
    if snapshot.insight:
        rows.append(("Headline", snapshot.insight.headline))
    for label, value in rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = _BOLD

    ws.append([])
    if snapshot.insight:
        ws.append(["Narrative"])
        ws.cell(row=ws.max_row, column=1).font = _BOLD
        ws.append([snapshot.insight.narrative])

        if snapshot.insight.claims:
            ws.append([])
            ws.append(["#", "Claim", "Evidence"])
            for cell in ws[ws.max_row]:
                cell.font = _BOLD
            for index, claim in enumerate(snapshot.insight.claims, start=1):
                ws.append([index, claim.text, ", ".join(claim.evidence)])
    elif snapshot.insight_error:
        ws.append(["Narrative unavailable"])
        ws.cell(row=ws.max_row, column=1).font = _BOLD
        ws.append([snapshot.insight_error])

    _autosize(ws)


def _write_data_sheet(ws: Worksheet, latest: AttemptRecord) -> None:
    ws.title = "Data"
    result = latest.validator.result
    assert result is not None  # guaranteed by app.action.policy before this is ever called

    ws.append(result.columns)
    for cell in ws[1]:
        cell.font = _BOLD
    for row in result.rows:
        ws.append(row)
    if result.truncated:
        ws.append([])
        ws.append([f"Truncated at {result.row_count} row(s); more rows exist in the warehouse."])

    _autosize(ws)


def _write_sql_evidence_sheet(ws: Worksheet, snapshot: RunSnapshot, latest: AttemptRecord) -> None:
    ws.title = "SQL & Evidence"
    ws.append(["SQL"])
    ws.cell(row=ws.max_row, column=1).font = _BOLD
    ws.append([latest.nl2sql.sql])

    if latest.nl2sql.parameters:
        ws.append([])
        ws.append(["Parameter", "Value"])
        for cell in ws[ws.max_row]:
            cell.font = _BOLD
        for key, value in latest.nl2sql.parameters.items():
            ws.append([key, value])

    if latest.nl2sql.assumptions:
        ws.append([])
        ws.append(["Assumptions"])
        ws.cell(row=ws.max_row, column=1).font = _BOLD
        for assumption in latest.nl2sql.assumptions:
            ws.append([assumption])

    if snapshot.retrieved_context:
        ws.append([])
        ws.append(["Retrieved definition", "Title", "Citation"])
        for cell in ws[ws.max_row]:
            cell.font = _BOLD
        for item in snapshot.retrieved_context:
            ws.append([item.object_name, item.title, item.citation])

    _autosize(ws)


def _write_validation_sheet(ws: Worksheet, snapshot: RunSnapshot) -> None:
    ws.title = "Validation"
    ws.append(["Attempt", "Check", "Status", "Details"])
    for cell in ws[1]:
        cell.font = _BOLD
    for attempt in snapshot.attempts:
        for check in attempt.validator.checks:
            ws.append([attempt.attempt_no, check.name, check.status, check.details])

    _autosize(ws)


def build_workbook(snapshot: RunSnapshot) -> bytes:
    """Builds the four-sheet workbook for an already-policy-checked,
    validated run. Callers (app.api.actions) must call
    app.action.policy.evaluate_action_policy first -- this function trusts
    that snapshot.attempts is non-empty and its last entry's validator
    passed with a result, the same way app.validator.executor trusts
    already-policy-checked SQL."""
    latest = snapshot.attempts[-1]

    workbook = Workbook()
    summary_ws = workbook.active
    assert summary_ws is not None
    _write_summary_sheet(summary_ws, snapshot, latest)
    _write_data_sheet(workbook.create_sheet(), latest)
    _write_sql_evidence_sheet(workbook.create_sheet(), snapshot, latest)
    _write_validation_sheet(workbook.create_sheet(), snapshot)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
