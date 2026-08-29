"""app.action.workbook.build_workbook -- builds real bytes, then reads
them back with openpyxl to verify sheet structure and content (the same
library, exercised as both writer and reader, is the simplest reliable
way to check a generated .xlsx file's actual content)."""

from __future__ import annotations

import datetime as dt
import uuid
from io import BytesIO

from openpyxl import load_workbook

from app.action.workbook import build_workbook
from app.catalog.schema import RetrievalResult
from app.insight.schema import Claim, InsightOutput
from app.nl2sql.schema import NL2SQLOutput
from app.orchestrator.schema import AttemptRecord, RunSnapshot
from app.validator.schema import QueryResult, ValidationCheck, ValidatorOutput

_NOW = dt.datetime.now(dt.UTC)


def _snapshot() -> RunSnapshot:
    result = QueryResult(
        columns=["quarter", "median_hold_hrs"],
        rows=[["2026-Q1", 9.5], ["2026-Q2", 27.4]],
        row_count=2,
        truncated=False,
    )
    nl2sql = NL2SQLOutput(
        sql="SELECT quarter, median_hold_hrs FROM analytics.v_task_lifecycle",
        dialect="postgres",
        referenced_objects=["analytics.v_task_lifecycle"],
        assumptions=["Q2 refers to the latest complete calendar quarter"],
        parameters={"department": "Buyer"},
        confidence=0.9,
    )
    failing_validator = ValidatorOutput(
        status="fail",
        checks=[ValidationCheck(name="sql_policy", status="fail", details="bad table")],
        repairable=True,
        feedback="bad table",
    )
    passing_validator = ValidatorOutput(
        status="pass",
        checks=[
            ValidationCheck(name="sql_policy", status="pass", details="ok"),
            ValidationCheck(name="result_not_empty", status="pass", details="2 row(s) returned."),
        ],
        repairable=False,
        result=result,
    )
    insight = InsightOutput(
        headline="Buyer department median hold time rose in Q2",
        narrative="Driven by Supplier Compliance Review tasks.",
        claims=[
            Claim(
                text="Median hold time moved from 9.5 hours to 27.4 hours",
                evidence=["result:r1:c2", "result:r2:c2"],
            )
        ],
    )
    return RunSnapshot(
        run_id=uuid.uuid4(),
        tenant_id="default",
        source_id="marketplace_demo",
        question="Why did median task hold time spike for the Buyer department in Q2?",
        status="READY",
        retrieved_context=[
            RetrievalResult(
                chunk_id=1,
                document_id=1,
                kind="table",
                object_name="analytics.v_task_lifecycle",
                title="View: analytics.v_task_lifecycle",
                content="one row per task",
                score=0.9,
                citation="catalog:table:analytics.v_task_lifecycle:chunk:0",
            )
        ],
        attempts=[
            AttemptRecord(attempt_no=1, nl2sql=nl2sql, validator=failing_validator),
            AttemptRecord(attempt_no=2, nl2sql=nl2sql, validator=passing_validator),
        ],
        insight=insight,
        created_at=_NOW,
        updated_at=_NOW,
        completed_at=_NOW,
    )


def test_workbook_has_the_four_expected_sheets_in_order() -> None:
    workbook = load_workbook(BytesIO(build_workbook(_snapshot())))

    assert workbook.sheetnames == ["Summary", "Data", "SQL & Evidence", "Validation"]


def test_summary_sheet_includes_the_question_and_headline() -> None:
    workbook = load_workbook(BytesIO(build_workbook(_snapshot())))
    values = {cell.value for row in workbook["Summary"].iter_rows() for cell in row}

    assert "Why did median task hold time spike for the Buyer department in Q2?" in values
    assert "Buyer department median hold time rose in Q2" in values
    assert "Median hold time moved from 9.5 hours to 27.4 hours" in values


def test_data_sheet_has_the_header_row_and_every_result_row() -> None:
    workbook = load_workbook(BytesIO(build_workbook(_snapshot())))
    ws = workbook["Data"]
    rows = list(ws.iter_rows(values_only=True))

    assert rows[0] == ("quarter", "median_hold_hrs")
    assert rows[1] == ("2026-Q1", 9.5)
    assert rows[2] == ("2026-Q2", 27.4)


def test_sql_and_evidence_sheet_includes_the_winning_sql_and_parameters() -> None:
    workbook = load_workbook(BytesIO(build_workbook(_snapshot())))
    values = [cell.value for row in workbook["SQL & Evidence"].iter_rows() for cell in row]

    assert "SELECT quarter, median_hold_hrs FROM analytics.v_task_lifecycle" in values
    assert "department" in values
    assert "Buyer" in values
    assert "analytics.v_task_lifecycle" in values


def test_validation_sheet_includes_checks_from_every_attempt() -> None:
    workbook = load_workbook(BytesIO(build_workbook(_snapshot())))
    rows = list(workbook["Validation"].iter_rows(values_only=True))

    attempt_numbers = {row[0] for row in rows[1:]}
    assert attempt_numbers == {1, 2}
    details = {row[3] for row in rows[1:]}
    assert "bad table" in details
    assert "ok" in details
