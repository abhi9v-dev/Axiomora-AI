"""FastAPI route tests for app.api.actions: status codes, idempotent
replay, policy rejection and error mapping -- all without a real database.

app.orchestrator.store.get_run and app.action.store's functions are
monkeypatched at the app.api.actions import site, backed by an in-memory
fake (same pattern as test_api_runs.py). app.action.workbook.build_workbook
runs for real -- it's pure, fast, and already unit-tested on its own in
test_workbook.py -- so these tests also verify a real .xlsx comes back.
"""

from __future__ import annotations

import contextlib
import dataclasses
import datetime as dt
import uuid
from collections.abc import Iterator
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

import app.api.actions as actions_module
from app.action.power_bi.base import PowerBIPushResult, PowerBIRefreshResult
from app.action.power_bi.errors import PowerBIAdapterError
from app.action.power_bi.mock import MockPowerBIAdapter
from app.action.schema import ActionRecord
from app.api.deps import get_power_bi_adapter_dep, get_session, get_settings_dep
from app.config import Settings
from app.insight.schema import InsightOutput
from app.main import create_app
from app.nl2sql.schema import NL2SQLOutput
from app.orchestrator.schema import AttemptRecord, RunSnapshot
from app.validator.schema import QueryResult, ValidationCheck, ValidatorOutput


def _settings(*, power_bi_enabled: bool = False) -> Settings:
    return Settings(
        _env_file=None,
        DATABASE_URL="postgresql+asyncpg://u:p@localhost:5432/app",
        WAREHOUSE_URL="postgresql+asyncpg://u:p@localhost:5432/wh",
        POWER_BI_ENABLED=power_bi_enabled,
    )


@dataclasses.dataclass
class _FlakyPowerBIAdapter:
    """Fails the first `fail_first_n` calls to either method, then
    succeeds -- used to exercise the failed-then-retried-and-recovered
    path without rebuilding the TestClient mid-test."""

    fail_first_n: int
    calls: int = 0

    async def push_rows(
        self, *, dataset_id: str, table_name: str, rows: list[dict[str, object]]
    ) -> PowerBIPushResult:
        self.calls += 1
        if self.calls <= self.fail_first_n:
            raise PowerBIAdapterError("simulated transient Power BI failure")
        return PowerBIPushResult(
            dataset_id=dataset_id, table_name=table_name, rows_pushed=len(rows)
        )

    async def refresh_dataset(self, *, dataset_id: str) -> PowerBIRefreshResult:
        self.calls += 1
        if self.calls <= self.fail_first_n:
            raise PowerBIAdapterError("simulated transient Power BI failure")
        return PowerBIRefreshResult(dataset_id=dataset_id, refresh_request_id="refresh-1")


def _snapshot(run_id: uuid.UUID, *, status: str = "READY", ready: bool = True) -> RunSnapshot:
    now = dt.datetime.now(dt.UTC)
    attempts = []
    if ready:
        attempts = [
            AttemptRecord(
                attempt_no=1,
                nl2sql=NL2SQLOutput(sql="SELECT 1", dialect="postgres", confidence=0.9),
                validator=ValidatorOutput(
                    status="pass",
                    checks=[ValidationCheck(name="sql_policy", status="pass", details="ok")],
                    repairable=False,
                    result=QueryResult(columns=["x"], rows=[[1]], row_count=1, truncated=False),
                ),
            )
        ]
    return RunSnapshot(
        run_id=run_id,
        tenant_id="default",
        source_id="marketplace_demo",
        question="Why did hold time spike?",
        status=status,
        attempts=attempts,
        insight=InsightOutput(headline="h", narrative="n") if ready else None,
        created_at=now,
        updated_at=now,
        completed_at=now if ready else None,
    )


class _FakeBackend:
    def __init__(self) -> None:
        self.runs: dict[uuid.UUID, RunSnapshot] = {}
        self.actions: dict[tuple[uuid.UUID, str], ActionRecord] = {}
        self.actions_by_id: dict[uuid.UUID, tuple[uuid.UUID, str]] = {}
        self.record_calls = 0
        self.update_calls = 0

    async def get_run(self, session: object, run_id: uuid.UUID) -> RunSnapshot | None:
        return self.runs.get(run_id)

    async def get_action_by_idempotency_key(
        self, session: object, run_id: uuid.UUID, idempotency_key: str
    ) -> ActionRecord | None:
        return self.actions.get((run_id, idempotency_key))

    async def record_action(self, session: object, record: ActionRecord) -> ActionRecord:
        self.record_calls += 1
        key = (record.run_id, record.idempotency_key)
        if key in self.actions:
            raise IntegrityError("duplicate", {}, Exception("uq_action_run_idempotency_key"))
        self.actions[key] = record
        self.actions_by_id[record.id] = key
        return record

    async def update_action_outcome(
        self, session: object, action_id: uuid.UUID, *, status: str, rejection_reason: str | None
    ) -> ActionRecord:
        self.update_calls += 1
        key = self.actions_by_id[action_id]
        updated = self.actions[key].model_copy(
            update={"status": status, "rejection_reason": rejection_reason}
        )
        self.actions[key] = updated
        return updated


@pytest.fixture
def backend(monkeypatch: pytest.MonkeyPatch) -> _FakeBackend:
    fake = _FakeBackend()
    monkeypatch.setattr(actions_module, "get_run", fake.get_run)
    monkeypatch.setattr(
        actions_module, "get_action_by_idempotency_key", fake.get_action_by_idempotency_key
    )
    monkeypatch.setattr(actions_module, "record_action", fake.record_action)
    monkeypatch.setattr(actions_module, "update_action_outcome", fake.update_action_outcome)
    return fake


class _FakeSession:
    """Stands in for AsyncSession -- the fake store functions never touch
    it for real queries, but app.api.actions calls session.rollback()
    after a simulated IntegrityError, so it needs to support that much."""

    async def rollback(self) -> None:
        return None


async def _fake_session() -> object:
    yield _FakeSession()


@pytest.fixture
def client(backend: _FakeBackend) -> Iterator[TestClient]:
    application: FastAPI = create_app()
    application.dependency_overrides[get_session] = _fake_session
    with TestClient(application) as test_client:
        yield test_client


@contextlib.contextmanager
def _make_client(
    backend: _FakeBackend,
    *,
    settings: Settings | None = None,
    power_bi_adapter: object | None = None,
) -> Iterator[TestClient]:
    application: FastAPI = create_app()
    application.dependency_overrides[get_session] = _fake_session
    if settings is not None:
        application.dependency_overrides[get_settings_dep] = lambda: settings
    if power_bi_adapter is not None:
        application.dependency_overrides[get_power_bi_adapter_dep] = lambda: power_bi_adapter
    with TestClient(application) as test_client:
        yield test_client


def test_export_excel_returns_a_real_workbook_for_a_ready_run(
    client: TestClient, backend: _FakeBackend
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)

    response = client.post(
        f"/api/v1/runs/{run_id}/actions",
        json={"type": "export_excel", "idempotency_key": "key-1"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in response.headers["content-disposition"]
    assert "X-Action-Id" in response.headers

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Data", "SQL & Evidence", "Validation"]
    assert backend.record_calls == 1


def test_repeating_the_same_idempotency_key_does_not_record_a_second_action(
    client: TestClient, backend: _FakeBackend
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)
    body = {"type": "export_excel", "idempotency_key": "same-key"}

    first = client.post(f"/api/v1/runs/{run_id}/actions", json=body)
    second = client.post(f"/api/v1/runs/{run_id}/actions", json=body)

    assert first.status_code == 200
    assert second.status_code == 200
    assert backend.record_calls == 1
    assert first.headers["X-Action-Id"] == second.headers["X-Action-Id"]


def test_a_power_bi_action_is_rejected_and_recorded_when_the_flag_is_disabled(
    client: TestClient, backend: _FakeBackend
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)

    response = client.post(
        f"/api/v1/runs/{run_id}/actions",
        json={"type": "power_bi_push", "idempotency_key": "key-1"},
    )

    assert response.status_code == 403
    assert "disabled" in response.json()["detail"]
    recorded = backend.actions[(run_id, "key-1")]
    assert recorded.status == "rejected"


def test_power_bi_replace_is_always_rejected_regardless_of_the_flag(backend: _FakeBackend) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)

    with _make_client(backend, settings=_settings(power_bi_enabled=True)) as test_client:
        response = test_client.post(
            f"/api/v1/runs/{run_id}/actions",
            json={"type": "power_bi_replace", "idempotency_key": "key-1"},
        )

    assert response.status_code == 403
    assert "prohibited" in response.json()["detail"]


def test_power_bi_push_succeeds_against_the_mock_adapter_once_enabled(
    backend: _FakeBackend,
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)
    adapter = MockPowerBIAdapter()

    with _make_client(
        backend, settings=_settings(power_bi_enabled=True), power_bi_adapter=adapter
    ) as test_client:
        response = test_client.post(
            f"/api/v1/runs/{run_id}/actions",
            json={"type": "power_bi_push", "idempotency_key": "key-1"},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "completed"
    assert body["destination"] == "power_bi:push"
    assert "Pushed 1 row" in body["detail"]
    assert response.headers["X-Action-Id"] == body["action_id"]
    assert len(adapter.pushed_calls) == 1
    dataset_id, table_name, rows = adapter.pushed_calls[0]
    assert dataset_id == "demo-dataset"
    assert table_name == "BiCopilotInsights"
    assert rows == [{"x": 1}]
    assert backend.actions[(run_id, "key-1")].status == "completed"


def test_power_bi_refresh_succeeds_against_the_mock_adapter_once_enabled(
    backend: _FakeBackend,
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)
    adapter = MockPowerBIAdapter()

    with _make_client(
        backend, settings=_settings(power_bi_enabled=True), power_bi_adapter=adapter
    ) as test_client:
        response = test_client.post(
            f"/api/v1/runs/{run_id}/actions",
            json={"type": "power_bi_refresh", "idempotency_key": "key-1"},
        )

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert response.json()["destination"] == "power_bi:refresh"
    assert adapter.refresh_calls == ["demo-dataset"]


def test_repeating_a_completed_power_bi_action_does_not_call_the_adapter_again(
    backend: _FakeBackend,
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)
    adapter = MockPowerBIAdapter()
    body = {"type": "power_bi_push", "idempotency_key": "same-key"}

    with _make_client(
        backend, settings=_settings(power_bi_enabled=True), power_bi_adapter=adapter
    ) as test_client:
        first = test_client.post(f"/api/v1/runs/{run_id}/actions", json=body)
        second = test_client.post(f"/api/v1/runs/{run_id}/actions", json=body)

    assert first.status_code == 200
    assert second.status_code == 200
    assert first.json()["action_id"] == second.json()["action_id"]
    assert first.json()["detail"] == second.json()["detail"]
    # The whole point of the idempotency key: a real external side effect
    # (pushing rows) must happen at most once per logical request, even
    # though the HTTP request was made twice.
    assert len(adapter.pushed_calls) == 1


def test_a_failed_power_bi_action_returns_502_and_a_retry_with_the_same_key_can_succeed(
    backend: _FakeBackend,
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)
    adapter = _FlakyPowerBIAdapter(fail_first_n=1)
    body = {"type": "power_bi_push", "idempotency_key": "same-key"}

    with _make_client(
        backend, settings=_settings(power_bi_enabled=True), power_bi_adapter=adapter
    ) as test_client:
        first = test_client.post(f"/api/v1/runs/{run_id}/actions", json=body)
        assert backend.actions[(run_id, "same-key")].status == "failed"

        second = test_client.post(f"/api/v1/runs/{run_id}/actions", json=body)

    assert first.status_code == 502
    assert "simulated transient" in first.json()["detail"]
    assert second.status_code == 200
    assert second.json()["status"] == "completed"
    # Retrying a failed action updates the same row rather than inserting
    # a new one (docs/03_ARCHITECTURE.md: "action failure: retain
    # validated answer and an idempotency key for safe retry").
    assert backend.record_calls == 1
    assert backend.update_calls == 2
    assert backend.actions[(run_id, "same-key")].status == "completed"


def test_power_bi_push_is_rejected_when_the_run_is_not_ready(backend: _FakeBackend) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id, status="GENERATING_SQL", ready=False)

    with _make_client(backend, settings=_settings(power_bi_enabled=True)) as test_client:
        response = test_client.post(
            f"/api/v1/runs/{run_id}/actions",
            json={"type": "power_bi_push", "idempotency_key": "key-1"},
        )

    assert response.status_code == 403


def test_a_repeated_rejected_request_stays_rejected_without_recording_again(
    client: TestClient, backend: _FakeBackend
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)
    body = {"type": "power_bi_push", "idempotency_key": "key-1"}

    client.post(f"/api/v1/runs/{run_id}/actions", json=body)
    second = client.post(f"/api/v1/runs/{run_id}/actions", json=body)

    assert second.status_code == 403
    assert backend.record_calls == 1


def test_export_is_rejected_when_the_run_is_not_ready(
    client: TestClient, backend: _FakeBackend
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id, status="GENERATING_SQL", ready=False)

    response = client.post(
        f"/api/v1/runs/{run_id}/actions",
        json={"type": "export_excel", "idempotency_key": "key-1"},
    )

    assert response.status_code == 403


def test_unknown_run_returns_404(client: TestClient) -> None:
    response = client.post(
        f"/api/v1/runs/{uuid.uuid4()}/actions",
        json={"type": "export_excel", "idempotency_key": "key-1"},
    )

    assert response.status_code == 404


async def test_a_concurrent_duplicate_insert_recovers_via_the_winning_row(
    client: TestClient, backend: _FakeBackend, monkeypatch: pytest.MonkeyPatch
) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)

    # Simulate two requests racing on the same idempotency key: this
    # request's own pre-check sees nothing (as if it ran before the other
    # request's commit), but by the time record_action actually runs, the
    # other request has already committed the row -- the endpoint must
    # recover by re-fetching that row rather than raising a 500.
    existing = ActionRecord(
        id=uuid.uuid4(),
        run_id=run_id,
        type="export_excel",
        destination="download",
        status="completed",
        idempotency_key="key-1",
        approved_by="result_owner",
        rejection_reason=None,
        created_at=dt.datetime.now(dt.UTC),
    )
    backend.actions[(run_id, "key-1")] = existing
    calls: list[int] = []

    async def _get_returns_none_once(
        session: object, r: uuid.UUID, key: str
    ) -> ActionRecord | None:
        if not calls:
            calls.append(1)
            return None
        return backend.actions.get((r, key))

    monkeypatch.setattr(actions_module, "get_action_by_idempotency_key", _get_returns_none_once)

    response = client.post(
        f"/api/v1/runs/{run_id}/actions",
        json={"type": "export_excel", "idempotency_key": "key-1"},
    )

    assert response.status_code == 200
    assert response.headers["X-Action-Id"] == str(existing.id)


def test_blank_idempotency_key_is_rejected(client: TestClient, backend: _FakeBackend) -> None:
    run_id = uuid.uuid4()
    backend.runs[run_id] = _snapshot(run_id)

    response = client.post(
        f"/api/v1/runs/{run_id}/actions",
        json={"type": "export_excel", "idempotency_key": "   "},
    )

    assert response.status_code == 422
